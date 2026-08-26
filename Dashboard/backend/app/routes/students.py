import uuid
from io import BytesIO
from typing import List

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from openpyxl import load_workbook
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.database.session import get_db
from app.deps.tenant import get_tenant_id
from app.models import Enrollment, Student, SubjectClass
from app.models.enums import EnrollmentStatus
from app.schemas.schemas import (
    EnrollmentCreate,
    EnrollmentSummary,
    StudentCreate,
    StudentDetailResponse,
    StudentResponse,
    StudentUpdate,
    StudentsListResponse,
    StudentHumanModeUpdate,
)
from app.services.dashboard_service import enrich_student, student_enrollment_summaries

router = APIRouter(prefix="/students", tags=["Students"])


@router.get("", response_model=StudentsListResponse)
def get_students(
    tenant_id: str = Depends(get_tenant_id),
    db: Session = Depends(get_db),
):
    students = (
        db.query(Student)
        .filter(Student.tenant_id == tenant_id)
        .order_by(Student.created_at.desc())
        .all()
    )

    enrollment_map = student_enrollment_summaries(db, [student.id for student in students])

    return {
        "tenant_id": tenant_id,
        "students": [
            {
                **StudentResponse.model_validate(student).model_dump(),
                "enrollments": enrollment_map.get(student.id, []),
            }
            for student in students
        ],
    }


@router.get("/by-phone/{phone}", response_model=StudentDetailResponse)
def get_student_by_phone(
    phone: str,
    tenant_id: str = Depends(get_tenant_id),
    db: Session = Depends(get_db),
):
    student = (
        db.query(Student)
        .filter(Student.tenant_id == tenant_id, Student.phone == phone)
        .first()
    )

    if student is None:
        raise HTTPException(status_code=404, detail="Student not found")

    return enrich_student(db, student)


@router.post("/import")
async def import_students(
    file: UploadFile = File(...),
    tenant_id: str = Depends(get_tenant_id),
    db: Session = Depends(get_db),
):
    if not (file.filename and file.filename.lower().endswith(".xlsx")):
        raise HTTPException(
            status_code=422,
            detail="Only .xlsx files are supported.",
        )

    content = await file.read()

    workbook = load_workbook(
        BytesIO(content),
        read_only=True,
        data_only=True,
    )

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    if not rows:
        raise HTTPException(status_code=422, detail="Excel file is empty.")

    headers = [str(value or "").strip().lower() for value in rows[0]]
    required = {"name", "phone"}

    if not required.issubset(set(headers)):
        raise HTTPException(
            status_code=422,
            detail="Excel must include name and phone columns.",
        )

    created = 0
    skipped = 0
    errors = []

    for row_number, values in enumerate(rows[1:], start=2):
        data = dict(zip(headers, values))

        name = str(data.get("name") or "").strip()
        phone = str(data.get("phone") or "").strip()

        if not phone:
            errors.append({"row": row_number, "reason": "Missing phone"})
            continue

        existing = (
            db.query(Student)
            .filter(
                Student.tenant_id == tenant_id,
                Student.phone == phone,
            )
            .first()
        )

        if existing:
            skipped += 1
            continue

        student = Student(
            id=str(uuid.uuid4()),
            tenant_id=tenant_id,
            name=name or None,
            phone=phone,
            district=(str(data.get("district") or "").strip() or None),
            language_pref=(str(data.get("language_pref") or "en").strip() or "en"),
        )

        db.add(student)
        db.flush()

        class_id = str(data.get("class_id") or "").strip()

        if class_id:
            subject_class = (
                db.query(SubjectClass)
                .filter(
                    SubjectClass.id == class_id,
                    SubjectClass.tenant_id == tenant_id,
                )
                .first()
            )

            if subject_class:
                db.add(
                    Enrollment(
                        id=str(uuid.uuid4()),
                        tenant_id=tenant_id,
                        student_id=student.id,
                        class_id=class_id,
                        status=EnrollmentStatus.PENDING,
                    )
                )

        created += 1

    db.commit()

    return {
        "created": created,
        "skipped": skipped,
        "errors": errors,
    }


@router.get("/{student_id}", response_model=StudentDetailResponse)
def get_student(
    student_id: str,
    tenant_id: str = Depends(get_tenant_id),
    db: Session = Depends(get_db),
):
    student = (
        db.query(Student)
        .filter(Student.id == student_id, Student.tenant_id == tenant_id)
        .first()
    )

    if student is None:
        raise HTTPException(status_code=404, detail="Student not found")

    return enrich_student(db, student)


@router.post("", response_model=StudentDetailResponse, status_code=201)
def create_student(
    student_data: StudentCreate,
    tenant_id: str = Depends(get_tenant_id),
    db: Session = Depends(get_db),
):
    if student_data.tenant_id != tenant_id:
        raise HTTPException(status_code=403, detail="tenant_id mismatch")

    if student_data.class_id:
        subject_class = (
            db.query(SubjectClass)
            .filter(
                SubjectClass.id == student_data.class_id,
                SubjectClass.tenant_id == tenant_id,
            )
            .first()
        )
        if subject_class is None:
            raise HTTPException(status_code=404, detail="Class not found")

    new_student = Student(
        id=str(uuid.uuid4()),
        tenant_id=tenant_id,
        name=student_data.name,
        phone=student_data.phone.strip(),
        district=student_data.district,
        language_pref=student_data.language_pref,
    )

    try:
        db.add(new_student)
        db.flush()

        if student_data.class_id:
            db.add(
                Enrollment(
                    id=str(uuid.uuid4()),
                    tenant_id=tenant_id,
                    student_id=new_student.id,
                    class_id=student_data.class_id,
                    status=EnrollmentStatus.PENDING,
                )
            )

        db.commit()
        db.refresh(new_student)
    except IntegrityError as error:
        db.rollback()
        raise HTTPException(
            status_code=409,
            detail="A student with this phone already exists for the tenant",
        ) from error

    return enrich_student(db, new_student)


@router.put("/{student_id}", response_model=StudentDetailResponse)
def update_student(
    student_id: str,
    student_data: StudentUpdate,
    tenant_id: str = Depends(get_tenant_id),
    db: Session = Depends(get_db),
):
    student = (
        db.query(Student)
        .filter(Student.id == student_id, Student.tenant_id == tenant_id)
        .first()
    )

    if student is None:
        raise HTTPException(status_code=404, detail="Student not found")

    if student_data.name is not None:
        student.name = student_data.name  # type: ignore[assignment]
    if student_data.phone is not None:
        student.phone = student_data.phone.strip()  # type: ignore[assignment]
    if student_data.district is not None:
        student.district = student_data.district  # type: ignore[assignment]
    if student_data.language_pref is not None:
        student.language_pref = student_data.language_pref  # type: ignore[assignment]

    try:
        db.commit()
        db.refresh(student)
    except IntegrityError as error:
        db.rollback()
        raise HTTPException(
            status_code=409,
            detail="A student with this phone already exists for the tenant",
        ) from error

    return enrich_student(db, student)


@router.patch(
    "/{student_id}/human-mode",
    response_model=StudentDetailResponse,
)
def update_student_human_mode(
    student_id: str,
    payload: StudentHumanModeUpdate,
    tenant_id: str = Depends(
        get_tenant_id
    ),
    db: Session = Depends(
        get_db
    ),
):
    student = (
        db.query(Student)
        .filter(
            Student.id == student_id,
            Student.tenant_id
            == tenant_id,
        )
        .first()
    )

    if student is None:
        raise HTTPException(
            status_code=404,
            detail="Student not found",
        )

    student.human_mode = (
        payload.human_mode
    )

    db.commit()
    db.refresh(student)

    return enrich_student(
        db,
        student,
    )


@router.delete("/{student_id}", status_code=204)
def delete_student(
    student_id: str,
    tenant_id: str = Depends(get_tenant_id),
    db: Session = Depends(get_db),
):
    student = (
        db.query(Student)
        .filter(Student.id == student_id, Student.tenant_id == tenant_id)
        .first()
    )

    if student is None:
        raise HTTPException(status_code=404, detail="Student not found")

    db.delete(student)
    db.commit()


@router.post(
    "/{student_id}/enrollments",
    response_model=EnrollmentSummary,
    status_code=201,
)
def enroll_student(
    student_id: str,
    enrollment_data: EnrollmentCreate,
    tenant_id: str = Depends(get_tenant_id),
    db: Session = Depends(get_db),
):
    student = (
        db.query(Student)
        .filter(Student.id == student_id, Student.tenant_id == tenant_id)
        .first()
    )

    if student is None:
        raise HTTPException(status_code=404, detail="Student not found")

    subject_class = (
        db.query(SubjectClass)
        .filter(
            SubjectClass.id == enrollment_data.class_id,
            SubjectClass.tenant_id == tenant_id,
        )
        .first()
    )

    if subject_class is None:
        raise HTTPException(status_code=404, detail="Class not found")

    enrollment = Enrollment(
        id=str(uuid.uuid4()),
        tenant_id=tenant_id,
        student_id=student_id,
        class_id=enrollment_data.class_id,
        status=enrollment_data.status,
    )

    try:
        db.add(enrollment)
        db.commit()
        db.refresh(enrollment)
    except IntegrityError as error:
        db.rollback()
        raise HTTPException(
            status_code=409,
            detail="Student is already enrolled in this class",
        ) from error

    return {
        "id": enrollment.id,
        "class_id": enrollment.class_id,
        "class_subject": subject_class.subject,
        "class_name": subject_class.name,
        "status": enrollment.status,
        "created_at": enrollment.created_at,
    }
